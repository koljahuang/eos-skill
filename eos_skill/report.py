"""
Excel report generator for EOS scan results.
Produces a formatted .xlsx file with the 12-column schema.
"""

from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Column definitions: (header_en, header_cn, field_key, width)
COLUMNS = [
    ("#", "序号", None, 6),
    ("Account", "账号", "account", 16),
    ("Region", "区域", "region", 16),
    ("Cluster Name", "集群名称", "cluster_name", 30),
    ("Instance Name", "实例名称", "instance_name", 30),
    ("Engine", "引擎", "engine", 14),
    ("Resource Type", "资源类型", "resource_type", 14),
    ("Instance Type", "实例类型", "instance_type", 20),
    ("Engine Version", "引擎版本", "engine_version", 16),
    ("End of Support Date", "停止支持日期", "eol_date", 20),
    ("Extended Support Date", "延长支持日期", "extended_support", 20),
    ("Target Engine Version", "目标版本号", "target_version", 20),
    ("Upgrade Type", "更新类型", "upgrade_type", 16),
]

# Styles
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="Arial", size=10)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Conditional row fills
EXPIRED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red-ish
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")       # Green


def _get_row_fill(eol_date) -> PatternFill | None:
    """Determine row color based on EOS proximity."""
    if eol_date is None:
        return None
    today = date.today()
    if isinstance(eol_date, datetime):
        eol_date = eol_date.date()
    if eol_date <= today:
        return EXPIRED_FILL  # Already expired
    days_remaining = (eol_date - today).days
    if days_remaining <= 180:
        return WARNING_FILL  # Expiring within 6 months
    return OK_FILL


def generate_report(rows: list[dict], output_path: str) -> str:
    """
    Generate an Excel report from scan results.
    Returns the output file path.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "EOS Report"

    # Freeze top row
    ws.freeze_panes = "A3"

    # Write bilingual header (row 1: English, row 2: Chinese)
    for col_idx, (header_en, header_cn, _, width) in enumerate(COLUMNS, start=1):
        # English header
        cell_en = ws.cell(row=1, column=col_idx, value=header_en)
        cell_en.font = HEADER_FONT
        cell_en.fill = HEADER_FILL
        cell_en.alignment = HEADER_ALIGNMENT
        cell_en.border = THIN_BORDER

        # Chinese header
        cell_cn = ws.cell(row=2, column=col_idx, value=header_cn)
        cell_cn.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell_cn.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell_cn.alignment = HEADER_ALIGNMENT
        cell_cn.border = THIN_BORDER

        # Column width
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=3):
        row_fill = _get_row_fill(row_data.get("eol_date"))

        for col_idx, (_, _, field_key, _) in enumerate(COLUMNS, start=1):
            if field_key is None:
                # Sequence number
                value = row_idx - 2
            else:
                value = row_data.get(field_key, "")
                # Format date
                if isinstance(value, date):
                    value = value.strftime("%Y-%m-%d")
                if value is None:
                    value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill

    # Auto-filter
    if rows:
        last_col = get_column_letter(len(COLUMNS))
        last_row = len(rows) + 2
        ws.auto_filter.ref = f"A2:{last_col}{last_row}"

    # Add legend sheet
    _add_legend(wb)

    wb.save(output_path)
    return output_path


def _add_legend(wb: Workbook):
    """Add a legend sheet explaining color coding."""
    ws = wb.create_sheet("Legend")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    ws.cell(row=1, column=1, value="Color").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Meaning").font = Font(bold=True)

    legends = [
        (EXPIRED_FILL, "Already past End-of-Support date (expired)"),
        (WARNING_FILL, "End-of-Support within 6 months (warning)"),
        (OK_FILL, "End-of-Support more than 6 months away (ok)"),
    ]
    for idx, (fill, desc) in enumerate(legends, start=2):
        cell_a = ws.cell(row=idx, column=1, value="Sample")
        cell_a.fill = fill
        ws.cell(row=idx, column=2, value=desc)
